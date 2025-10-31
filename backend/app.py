import os
import time
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
import json
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Initialize Flask app
app = Flask(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure CORS
CORS(app)

# Database configuration
database_url = os.environ.get('DATABASE_URL', 'postgresql://pos_user:pos_password@localhost:5432/pos_db')

# Handle MySQL URL format (PythonAnywhere uses mysql:// but SQLAlchemy needs mysql+pymysql://)
if database_url.startswith('mysql://'):
    database_url = database_url.replace('mysql://', 'mysql+pymysql://')

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 280,  # Prevent MySQL connection timeout
    'pool_pre_ping': True,  # Verify connections before using
}

# Import models after db initialization
from models import db, Table, TableOrder, Invoice, KOTConfig, BillConfig, MenuItem, Category, Department, RestaurantSettings

# Initialize database
db.init_app(app)

# Retry database connection
def connect_db():
    retries = 5
    while retries > 0:
        try:
            with app.app_context():
                db.create_all()
            logger.info("Database connected successfully")
            return True
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            retries -= 1
            if retries == 0:
                raise e
            logger.info(f"Retrying in 5 seconds... ({retries} attempts left)")
            time.sleep(5)

# Connect to database
try:
    connect_db()
except Exception as e:
    logger.error(f"Failed to connect to database after retries: {e}")

# Routes
@app.route('/api/tables', methods=['GET'])
def get_tables():
    """Get all tables"""
    try:
        tables = Table.query.all()
        return jsonify([table.to_dict() for table in tables])
    except Exception as e:
        logger.error(f"Error getting tables: {e}")
        return jsonify({'error': 'Failed to retrieve tables'}), 500

@app.route('/api/tables', methods=['POST'])
def create_table():
    """Create a new table"""
    try:
        data = request.get_json()
        
        new_table = Table(
            id=data.get('id', str(int(time.time() * 1000))),  # Generate ID if not provided
            name=data['name'],
            seats=data['seats'],
            category=data['category'],
            status=data.get('status', 'available')
        )
        
        db.session.add(new_table)
        db.session.commit()
        
        return jsonify(new_table.to_dict()), 201
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        return jsonify({'error': 'Failed to create table'}), 500

@app.route('/api/tables/<string:table_id>', methods=['PUT'])
def update_table(table_id):
    """Update a table"""
    try:
        table = Table.query.get(table_id)
        if not table:
            return jsonify({'error': 'Table not found'}), 404
        
        data = request.get_json()
        table.name = data.get('name', table.name)
        table.seats = data.get('seats', table.seats)
        table.category = data.get('category', table.category)
        table.status = data.get('status', table.status)
        
        db.session.commit()
        
        return jsonify(table.to_dict())
    except Exception as e:
        logger.error(f"Error updating table: {e}")
        return jsonify({'error': 'Failed to update table'}), 500

@app.route('/api/tables/<string:table_id>', methods=['DELETE'])
def delete_table(table_id):
    """Delete a table"""
    try:
        table = Table.query.get(table_id)
        if not table:
            return jsonify({'error': 'Table not found'}), 404
        
        db.session.delete(table)
        db.session.commit()
        
        return jsonify({'message': 'Table deleted successfully'})
    except Exception as e:
        logger.error(f"Error deleting table: {e}")
        return jsonify({'error': 'Failed to delete table'}), 500

@app.route('/api/orders', methods=['GET'])
def get_orders():
    """Get all orders"""
    try:
        orders = TableOrder.query.all()
        return jsonify([order.to_dict() for order in orders])
    except Exception as e:
        logger.error(f"Error getting orders: {e}")
        return jsonify({'error': 'Failed to retrieve orders'}), 500

@app.route('/api/orders/table/<string:table_id>', methods=['GET'])
def get_table_order(table_id):
    """Get order for a specific table"""
    try:
        order = TableOrder.query.filter_by(table_id=table_id).first()
        if order:
            return jsonify(order.to_dict())
        return jsonify(None)
    except Exception as e:
        logger.error(f"Error getting table order: {e}")
        return jsonify({'error': 'Failed to retrieve table order'}), 500

@app.route('/api/orders/table/<string:table_id>', methods=['POST'])
def add_items_to_table(table_id):
    """Add items to a table order"""
    try:
        data = request.get_json()
        
        # Check if order exists for this table
        order = TableOrder.query.filter_by(table_id=table_id).first()
        
        if not order:
            # Create new order
            order = TableOrder(
                table_id=table_id,
                table_name=data['table_name'],
                items=json.dumps(data['items']),
                start_time=datetime.now()
            )
            db.session.add(order)
            
            # Update table status
            table = Table.query.get(table_id)
            if table:
                table.status = 'occupied'
        else:
            # Update existing order
            existing_items = json.loads(order.items) if order.items else []
            new_items = data['items']
            
            # Merge items
            for new_item in new_items:
                existing_item_index = None
                for i, item in enumerate(existing_items):
                    if item['id'] == new_item['id'] and item.get('sentToKitchen', False):
                        # Item already sent to kitchen, add as new item
                        existing_items.append(new_item)
                        existing_item_index = -1
                        break
                    elif item['id'] == new_item['id'] and not item.get('sentToKitchen', False):
                        # Update quantity of pending item
                        existing_items[i]['quantity'] += new_item['quantity']
                        existing_item_index = i
                        break
                
                if existing_item_index is None:
                    # New item
                    existing_items.append(new_item)
            
            order.items = json.dumps(existing_items)
            
            # Update table status
            table = Table.query.get(table_id)
            if table:
                table.status = 'occupied'
        
        db.session.commit()
        
        return jsonify(order.to_dict())
    except Exception as e:
        logger.error(f"Error adding items to table: {e}")
        return jsonify({'error': 'Failed to add items to table'}), 500

@app.route('/api/orders/table/<string:table_id>/sent', methods=['POST'])
def mark_items_as_sent(table_id):
    """Mark all items in an order as sent to kitchen"""
    try:
        order = TableOrder.query.filter_by(table_id=table_id).first()
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        items = json.loads(order.items) if order.items else []
        for item in items:
            item['sentToKitchen'] = True
        
        order.items = json.dumps(items)
        db.session.commit()
        
        return jsonify(order.to_dict())
    except Exception as e:
        logger.error(f"Error marking items as sent: {e}")
        return jsonify({'error': 'Failed to mark items as sent'}), 500

@app.route('/api/orders/table/<string:table_id>/complete', methods=['POST'])
def complete_table_order(table_id):
    """Complete an order and remove it"""
    try:
        order = TableOrder.query.filter_by(table_id=table_id).first()
        if order:
            db.session.delete(order)
        
        # Update table status
        table = Table.query.get(table_id)
        if table:
            table.status = 'available'
        
        db.session.commit()
        
        return jsonify({'message': 'Order completed successfully'})
    except Exception as e:
        logger.error(f"Error completing table order: {e}")
        return jsonify({'error': 'Failed to complete order'}), 500

@app.route('/api/invoices', methods=['GET'])
def get_invoices():
    """Get all invoices"""
    try:
        invoices = Invoice.query.all()
        return jsonify([invoice.to_dict() for invoice in invoices])
    except Exception as e:
        logger.error(f"Error getting invoices: {e}")
        return jsonify({'error': 'Failed to retrieve invoices'}), 500

@app.route('/api/invoices', methods=['POST'])
def add_invoice():
    """Add a new invoice"""
    try:
        data = request.get_json()
        
        new_invoice = Invoice(
            id=data.get('id', str(int(time.time() * 1000))),  # Generate ID if not provided
            bill_number=data['billNumber'],
            order_type=data['orderType'],
            table_name=data.get('tableName'),
            items=json.dumps(data['items']),
            subtotal=data['subtotal'],
            tax=data['tax'],
            total=data['total'],
            timestamp=datetime.fromisoformat(data['timestamp'].replace('Z', '+00:00'))
        )
        
        db.session.add(new_invoice)
        db.session.commit()
        
        return jsonify(new_invoice.to_dict()), 201
    except Exception as e:
        logger.error(f"Error adding invoice: {e}")
        return jsonify({'error': 'Failed to add invoice'}), 500

@app.route('/api/config/kot', methods=['GET'])
def get_kot_config():
    """Get KOT configuration"""
    try:
        config = KOTConfig.query.first()
        if not config:
            # Create default config
            config = KOTConfig(
                print_by_department=False,
                number_of_copies=1
            )
            db.session.add(config)
            db.session.commit()
        
        return jsonify(config.to_dict())
    except Exception as e:
        logger.error(f"Error getting KOT config: {e}")
        return jsonify({'error': 'Failed to retrieve KOT configuration'}), 500

@app.route('/api/config/kot', methods=['PUT'])
def update_kot_config():
    """Update KOT configuration"""
    try:
        config = KOTConfig.query.first()
        if not config:
            config = KOTConfig()
            db.session.add(config)
        
        data = request.get_json()
        config.print_by_department = data.get('printByDepartment', config.print_by_department)
        config.number_of_copies = data.get('numberOfCopies', config.number_of_copies)
        config.selected_printer = data.get('selectedPrinter', config.selected_printer)
        
        db.session.commit()
        
        return jsonify(config.to_dict())
    except Exception as e:
        logger.error(f"Error updating KOT config: {e}")
        return jsonify({'error': 'Failed to update KOT configuration'}), 500

@app.route('/api/config/bill', methods=['GET'])
def get_bill_config():
    """Get bill configuration"""
    try:
        config = BillConfig.query.first()
        if not config:
            # Create default config
            config = BillConfig(
                auto_print_dine_in=False,
                auto_print_takeaway=False
            )
            db.session.add(config)
            db.session.commit()
        
        return jsonify(config.to_dict())
    except Exception as e:
        logger.error(f"Error getting bill config: {e}")
        return jsonify({'error': 'Failed to retrieve bill configuration'}), 500

@app.route('/api/config/bill', methods=['PUT'])
def update_bill_config():
    """Update bill configuration"""
    try:
        config = BillConfig.query.first()
        if not config:
            config = BillConfig()
            db.session.add(config)
        
        data = request.get_json()
        config.auto_print_dine_in = data.get('autoPrintDineIn', config.auto_print_dine_in)
        config.auto_print_takeaway = data.get('autoPrintTakeaway', config.auto_print_takeaway)
        config.selected_printer = data.get('selectedPrinter', config.selected_printer)
        
        db.session.commit()
        
        return jsonify(config.to_dict())
    except Exception as e:
        logger.error(f"Error updating bill config: {e}")
        return jsonify({'error': 'Failed to update bill configuration'}), 500

@app.route('/api/login', methods=['POST'])
def login():
    """User login"""
    try:
        data = request.get_json()
        # In a real application, you would validate credentials against a user database
        # For now, we'll just return a success response
        return jsonify({'message': 'Login successful'})
    except Exception as e:
        logger.error(f"Error during login: {e}")
        return jsonify({'error': 'Login failed'}), 500

# Menu Item API
@app.route('/api/menu-items', methods=['GET'])
def get_menu_items():
    """Get all menu items"""
    try:
        items = MenuItem.query.all()
        return jsonify([item.to_dict() for item in items])
    except Exception as e:
        logger.error(f"Error getting menu items: {e}")
        return jsonify({'error': 'Failed to retrieve menu items'}), 500

@app.route('/api/menu-items', methods=['POST'])
def create_menu_item():
    """Create a new menu item"""
    try:
        data = request.get_json()
        
        # Check if product_code already exists
        existing = MenuItem.query.filter_by(product_code=data['productCode']).first()
        if existing:
            return jsonify({'error': 'Product code already exists'}), 400
        
        new_item = MenuItem(
            id=data.get('id', str(int(time.time() * 1000))),
            name=data['name'],
            product_code=data['productCode'],
            price=data['price'],
            category=data['category'],
            department=data['department'],
            description=data.get('description', '')
        )
        
        db.session.add(new_item)
        db.session.commit()
        
        return jsonify(new_item.to_dict()), 201
    except Exception as e:
        logger.error(f"Error creating menu item: {e}")
        return jsonify({'error': 'Failed to create menu item'}), 500

@app.route('/api/menu-items/<string:item_id>', methods=['PUT'])
def update_menu_item(item_id):
    """Update a menu item"""
    try:
        item = MenuItem.query.get(item_id)
        if not item:
            return jsonify({'error': 'Menu item not found'}), 404
        
        data = request.get_json()
        
        # Check if product_code is being changed and if new code already exists
        if 'productCode' in data and data['productCode'] != item.product_code:
            existing = MenuItem.query.filter_by(product_code=data['productCode']).first()
            if existing:
                return jsonify({'error': 'Product code already exists'}), 400
            item.product_code = data['productCode']
        
        item.name = data.get('name', item.name)
        item.price = data.get('price', item.price)
        item.category = data.get('category', item.category)
        item.department = data.get('department', item.department)
        item.description = data.get('description', item.description)
        
        db.session.commit()
        
        return jsonify(item.to_dict())
    except Exception as e:
        logger.error(f"Error updating menu item: {e}")
        return jsonify({'error': 'Failed to update menu item'}), 500

@app.route('/api/menu-items/<string:item_id>', methods=['DELETE'])
def delete_menu_item(item_id):
    """Delete a menu item"""
    try:
        item = MenuItem.query.get(item_id)
        if not item:
            return jsonify({'error': 'Menu item not found'}), 404
        
        db.session.delete(item)
        db.session.commit()
        
        return jsonify({'message': 'Menu item deleted successfully'})
    except Exception as e:
        logger.error(f"Error deleting menu item: {e}")
        return jsonify({'error': 'Failed to delete menu item'}), 500

# Category API
@app.route('/api/categories', methods=['GET'])
def get_categories():
    """Get all categories"""
    try:
        categories = Category.query.all()
        return jsonify([cat.to_dict() for cat in categories])
    except Exception as e:
        logger.error(f"Error getting categories: {e}")
        return jsonify({'error': 'Failed to retrieve categories'}), 500

@app.route('/api/categories', methods=['POST'])
def create_category():
    """Create a new category"""
    try:
        data = request.get_json()
        
        new_category = Category(
            id=data.get('id', str(int(time.time() * 1000))),
            name=data['name']
        )
        
        db.session.add(new_category)
        db.session.commit()
        
        return jsonify(new_category.to_dict()), 201
    except Exception as e:
        logger.error(f"Error creating category: {e}")
        return jsonify({'error': 'Failed to create category'}), 500

@app.route('/api/categories/<string:category_id>', methods=['DELETE'])
def delete_category(category_id):
    """Delete a category"""
    try:
        category = Category.query.get(category_id)
        if not category:
            return jsonify({'error': 'Category not found'}), 404
        
        db.session.delete(category)
        db.session.commit()
        
        return jsonify({'message': 'Category deleted successfully'})
    except Exception as e:
        logger.error(f"Error deleting category: {e}")
        return jsonify({'error': 'Failed to delete category'}), 500

# Department API
@app.route('/api/departments', methods=['GET'])
def get_departments():
    """Get all departments"""
    try:
        departments = Department.query.all()
        return jsonify([dept.to_dict() for dept in departments])
    except Exception as e:
        logger.error(f"Error getting departments: {e}")
        return jsonify({'error': 'Failed to retrieve departments'}), 500

@app.route('/api/departments', methods=['POST'])
def create_department():
    """Create a new department"""
    try:
        data = request.get_json()
        
        new_department = Department(
            id=data.get('id', str(int(time.time() * 1000))),
            name=data['name']
        )
        
        db.session.add(new_department)
        db.session.commit()
        
        return jsonify(new_department.to_dict()), 201
    except Exception as e:
        logger.error(f"Error creating department: {e}")
        return jsonify({'error': 'Failed to create department'}), 500

@app.route('/api/departments/<string:department_id>', methods=['DELETE'])
def delete_department(department_id):
    """Delete a department"""
    try:
        department = Department.query.get(department_id)
        if not department:
            return jsonify({'error': 'Department not found'}), 404
        
        db.session.delete(department)
        db.session.commit()
        
        return jsonify({'message': 'Department deleted successfully'})
    except Exception as e:
        logger.error(f"Error deleting department: {e}")
        return jsonify({'error': 'Failed to delete department'}), 500

# Restaurant Settings API
@app.route('/api/restaurant-settings', methods=['GET'])
def get_restaurant_settings():
    """Get restaurant settings"""
    try:
        settings = RestaurantSettings.query.first()
        if not settings:
            # Create default settings
            settings = RestaurantSettings(
                restaurant_name='My Restaurant',
                currency='INR',
                tax_rate=5.0
            )
            db.session.add(settings)
            db.session.commit()
        
        return jsonify(settings.to_dict())
    except Exception as e:
        logger.error(f"Error getting restaurant settings: {e}")
        return jsonify({'error': 'Failed to retrieve restaurant settings'}), 500

@app.route('/api/restaurant-settings', methods=['PUT'])
def update_restaurant_settings():
    """Update restaurant settings"""
    try:
        settings = RestaurantSettings.query.first()
        if not settings:
            settings = RestaurantSettings()
            db.session.add(settings)
        
        data = request.get_json()
        settings.restaurant_name = data.get('restaurantName', settings.restaurant_name)
        settings.address = data.get('address', settings.address)
        settings.phone = data.get('phone', settings.phone)
        settings.email = data.get('email', settings.email)
        settings.currency = data.get('currency', settings.currency)
        settings.tax_rate = data.get('taxRate', settings.tax_rate)
        
        db.session.commit()
        
        return jsonify(settings.to_dict())
    except Exception as e:
        logger.error(f"Error updating restaurant settings: {e}")
        return jsonify({'error': 'Failed to update restaurant settings'}), 500

# Excel Import/Export API
@app.route('/api/menu/export-template', methods=['GET'])
def export_menu_template():
    """Download Excel template for bulk menu import"""
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Create Categories sheet
        ws_categories = wb.active
        ws_categories.title = "Categories"
        ws_categories['A1'] = "Category Name"
        ws_categories['A1'].font = Font(bold=True)
        ws_categories['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_categories['A1'].font = Font(bold=True, color="FFFFFF")
        ws_categories.column_dimensions['A'].width = 30
        
        # Add sample data
        ws_categories['A2'] = "Example: Appetizers"
        
        # Create Departments sheet
        ws_departments = wb.create_sheet("Departments")
        ws_departments['A1'] = "Department Name"
        ws_departments['A1'].font = Font(bold=True)
        ws_departments['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_departments['A1'].font = Font(bold=True, color="FFFFFF")
        ws_departments.column_dimensions['A'].width = 30
        
        # Add sample data
        ws_departments['A2'] = "Example: Kitchen"
        
        # Create Menu Items sheet
        ws_items = wb.create_sheet("Menu Items")
        headers = ["Product Code", "Item Name", "Price", "Category", "Department", "Description"]
        for col, header in enumerate(headers, start=1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        ws_items.column_dimensions['A'].width = 20
        ws_items.column_dimensions['B'].width = 30
        ws_items.column_dimensions['C'].width = 15
        ws_items.column_dimensions['D'].width = 20
        ws_items.column_dimensions['E'].width = 20
        ws_items.column_dimensions['F'].width = 50
        
        # Add sample data
        ws_items['A2'] = "CB001"
        ws_items['B2'] = "Example: Chicken Burger"
        ws_items['C2'] = 299
        ws_items['D2'] = "Mains"
        ws_items['E2'] = "Kitchen"
        ws_items['F2'] = "Grilled chicken with lettuce and mayo"
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='menu_import_template.xlsx'
        )
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        return jsonify({'error': 'Failed to generate template'}), 500

@app.route('/api/menu/export', methods=['GET'])
def export_menu_data():
    """Export current menu data to Excel"""
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Export Categories
        ws_categories = wb.active
        ws_categories.title = "Categories"
        ws_categories['A1'] = "Category Name"
        ws_categories['A1'].font = Font(bold=True, color="FFFFFF")
        ws_categories['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_categories.column_dimensions['A'].width = 30
        
        categories = Category.query.all()
        for idx, cat in enumerate(categories, start=2):
            ws_categories[f'A{idx}'] = cat.name
        
        # Export Departments
        ws_departments = wb.create_sheet("Departments")
        ws_departments['A1'] = "Department Name"
        ws_departments['A1'].font = Font(bold=True, color="FFFFFF")
        ws_departments['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_departments.column_dimensions['A'].width = 30
        
        departments = Department.query.all()
        for idx, dept in enumerate(departments, start=2):
            ws_departments[f'A{idx}'] = dept.name
        
        # Export Menu Items
        ws_items = wb.create_sheet("Menu Items")
        headers = ["Product Code", "Item Name", "Price", "Category", "Department", "Description"]
        for col, header in enumerate(headers, start=1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        ws_items.column_dimensions['A'].width = 20
        ws_items.column_dimensions['B'].width = 30
        ws_items.column_dimensions['C'].width = 15
        ws_items.column_dimensions['D'].width = 20
        ws_items.column_dimensions['E'].width = 20
        ws_items.column_dimensions['F'].width = 50
        
        items = MenuItem.query.all()
        for idx, item in enumerate(items, start=2):
            ws_items[f'A{idx}'] = item.product_code
            ws_items[f'B{idx}'] = item.name
            ws_items[f'C{idx}'] = item.price
            ws_items[f'D{idx}'] = item.category
            ws_items[f'E{idx}'] = item.department
            ws_items[f'F{idx}'] = item.description
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='menu_data_export.xlsx'
        )
    except Exception as e:
        logger.error(f"Error exporting menu data: {e}")
        return jsonify({'error': 'Failed to export menu data'}), 500

@app.route('/api/menu/import', methods=['POST'])
def import_menu_data():
    """Import menu data from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400
        
        # Load workbook
        wb = load_workbook(file)
        
        stats = {
            'categories_added': 0,
            'departments_added': 0,
            'items_added': 0,
            'errors': []
        }
        
        # Import Categories
        if 'Categories' in wb.sheetnames:
            ws_categories = wb['Categories']
            for row in ws_categories.iter_rows(min_row=2, values_only=True):
                if row[0] and not str(row[0]).startswith('Example:'):
                    category_name = str(row[0]).strip()
                    # Check if category already exists
                    existing = Category.query.filter_by(name=category_name).first()
                    if not existing:
                        new_category = Category(
                            id=str(int(time.time() * 1000)) + str(stats['categories_added']),
                            name=category_name
                        )
                        db.session.add(new_category)
                        stats['categories_added'] += 1
        
        # Import Departments
        if 'Departments' in wb.sheetnames:
            ws_departments = wb['Departments']
            for row in ws_departments.iter_rows(min_row=2, values_only=True):
                if row[0] and not str(row[0]).startswith('Example:'):
                    dept_name = str(row[0]).strip()
                    # Check if department already exists
                    existing = Department.query.filter_by(name=dept_name).first()
                    if not existing:
                        new_dept = Department(
                            id=str(int(time.time() * 1000)) + str(stats['departments_added']),
                            name=dept_name
                        )
                        db.session.add(new_dept)
                        stats['departments_added'] += 1
        
        # Commit categories and departments first
        db.session.commit()
        
        # Import Menu Items
        if 'Menu Items' in wb.sheetnames:
            ws_items = wb['Menu Items']
            for row_idx, row in enumerate(ws_items.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if row[1] and not str(row[1]).startswith('Example:'):
                        product_code = str(row[0]).strip() if row[0] else ''
                        name = str(row[1]).strip()
                        price = float(row[2]) if row[2] else 0
                        category = str(row[3]).strip() if row[3] else ''
                        department = str(row[4]).strip() if row[4] else ''
                        description = str(row[5]).strip() if row[5] else ''
                        
                        # Validate required fields
                        if not product_code:
                            stats['errors'].append(f"Row {row_idx}: Product code is required")
                            continue
                        
                        # Check if product code already exists
                        if MenuItem.query.filter_by(product_code=product_code).first():
                            stats['errors'].append(f"Row {row_idx}: Product code '{product_code}' already exists")
                            continue
                        
                        # Validate category and department exist
                        if category and not Category.query.filter_by(name=category).first():
                            stats['errors'].append(f"Row {row_idx}: Category '{category}' does not exist")
                            continue
                        
                        if department and not Department.query.filter_by(name=department).first():
                            stats['errors'].append(f"Row {row_idx}: Department '{department}' does not exist")
                            continue
                        
                        new_item = MenuItem(
                            id=str(int(time.time() * 1000)) + str(stats['items_added']),
                            name=name,
                            product_code=product_code,
                            price=price,
                            category=category,
                            department=department,
                            description=description
                        )
                        db.session.add(new_item)
                        stats['items_added'] += 1
                except Exception as e:
                    stats['errors'].append(f"Row {row_idx}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Import completed successfully',
            'stats': stats
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error importing menu data: {e}")
        return jsonify({'error': f'Failed to import menu data: {str(e)}'}), 500

# Serve React App (for production deployment)
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react_app(path):
    """Serve the React frontend application"""
    try:
        # Get the build folder path (parent directory of backend)
        build_folder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'build')
        
        # If path is provided and file exists, serve it
        if path != "" and os.path.exists(os.path.join(build_folder, path)):
            return send_file(os.path.join(build_folder, path))
        
        # Check if it's an API request (should have been handled by API routes)
        if path.startswith('api/'):
            return jsonify({'error': 'API endpoint not found'}), 404
        
        # Otherwise, serve index.html for React routing
        index_path = os.path.join(build_folder, 'index.html')
        if os.path.exists(index_path):
            return send_file(index_path)
        else:
            return jsonify({'error': 'Frontend build not found. Please run: npm run build'}), 404
    except Exception as e:
        logger.error(f"Error serving frontend: {e}")
        return jsonify({'error': 'Failed to serve frontend application'}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)